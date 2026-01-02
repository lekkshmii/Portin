#!/usr/bin/env python3
"""
ULTRA-OPTIMIZED M&A SCREENER DEMO
Uses MAXIMUM free-tier services to cost $0

FREE SERVICES USED:
- Gemini 2.5 Flash: 250 req/day (PRIMARY LLM)
- Firecrawl: 500 credits (WEB SCRAPING)
- Apify: $5 monthly credits (BACKUP SCRAPING)
- Serper.dev: 2,500 searches/month (GOOGLE SEARCH)
- OpenCage: 2,500 req/day (GEOCODING)
- requests: Local HTTP scraping

TOTAL COST: $0
"""

import os
import json
import time
import re
from datetime import datetime
from typing import Dict, List, Optional
import requests
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from config.model_config import get_current_model
# NEW: Import Deep Research Module
try:
    from enrichment.deep_research import DeepResearchEnricher, EnrichedCompany
    DEEP_RESEARCH_AVAILABLE = True
except ImportError as e:
    print(f"[WARNING] Deep Research module not found: {e}")
    DEEP_RESEARCH_AVAILABLE = False

load_dotenv()

# Configure Gemini
genai.configure(api_key=os.getenv('GEMINI_API_KEY'))
model = genai.GenerativeModel(get_current_model())

class MultiSourceEnricher:
    """
    Uses multiple free-tier services with smart fallbacks
    Priority: Free > Reliable > Fast
    """
    
    def __init__(self):
        self.serper_key = os.getenv('SERPER_KEY')
        self.firecrawl_key = os.getenv('FIRECRAWL_KEY')
        self.apify_key = os.getenv('APIFY_KEY', '')
        
        # Rate limiting
        self.last_gemini_call = 0
        self.gemini_min_interval = 6  # 10 req/min = 6 sec between calls
        
        # Scraper preference (set by user)
        self.scraper_choice = None  # 'crawl4ai', 'firecrawl', or 'both'
        
        # Stats
        self.stats = {
            'gemini_calls': 0,
            'grounding_calls': 0,
            'serper_calls': 0,
            'firecrawl_calls': 0,
            'crawl4ai_calls': 0,
            'http_calls': 0
        }
        
        # Deep Enricher lazy load or init
        self.deep_enricher = None
        if DEEP_RESEARCH_AVAILABLE:
            self.deep_enricher = DeepResearchEnricher()
    
    def ask_scraper_preference(self):
        """Ask user which scraper to use for enrichment."""
        print("\n" + "─"*50)
        print("SCRAPER SELECTION")
        print("─"*50)
        print("\n[1] Crawl4AI - Free, local browser (recommended)")
        print("[2] Firecrawl - API-based, 500 free credits")
        print("[3] Both - Crawl4AI first, Firecrawl fallback\n")
        
        while True:
            choice = input("Select scraper [1/2/3]: ").strip()
            
            if choice == "1":
                self.scraper_choice = "crawl4ai"
                print("\n[INFO] Using Crawl4AI (Free)\n")
                return
            elif choice == "2":
                self.scraper_choice = "firecrawl"
                print("\n[INFO] Using Firecrawl API\n")
                return
            elif choice == "3":
                self.scraper_choice = "both"
                print("\n[INFO] Using Crawl4AI with Firecrawl fallback\n")
                return
            else:
                print("[ERROR] Please enter 1, 2, or 3")
    
    def _wait_for_gemini_rate_limit(self):
        """Ensure we don't hit Gemini rate limits"""
        elapsed = time.time() - self.last_gemini_call
        if elapsed < self.gemini_min_interval:
            time.sleep(self.gemini_min_interval - elapsed)
        self.last_gemini_call = time.time()
    
    def search_company(self, company_name: str) -> Dict:
        """Step 1: Search for company using Serper (2,500 free searches)"""
        if not self.serper_key:
            print("[WARNING] No Serper key - using HTTP only")
            return {}
        
        try:
            url = "https://google.serper.dev/search"
            headers = {
                'X-API-KEY': self.serper_key,
                'Content-Type': 'application/json'
            }
            payload = {
                "q": f"{company_name} promotional products USA company",
                "num": 3
            }
            
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            self.stats['serper_calls'] += 1
            
            if response.status_code == 200:
                data = response.json()
                organic = data.get('organic', [])
                
                if organic:
                    result = organic[0]
                    return {
                        'website': result.get('link', ''),
                        'snippet': result.get('snippet', ''),
                        'title': result.get('title', '')
                    }
        except Exception as e:
            print(f"   [WARNING] Serper error: {e}")
        
        return {}
    
    def scrape_website_firecrawl(self, url: str) -> str:
        """Step 2: Scrape website using Firecrawl (500 free credits)"""
        if not self.firecrawl_key or not url:
            return ""
        
        try:
            api_url = "https://api.firecrawl.dev/v1/scrape"
            headers = {
                'Authorization': f'Bearer {self.firecrawl_key}',
                'Content-Type': 'application/json'
            }
            payload = {
                'url': url,
                'formats': ['markdown'],
                'onlyMainContent': True
            }
            
            response = requests.post(api_url, headers=headers, json=payload, timeout=30)
            self.stats['firecrawl_calls'] += 1
            
            if response.status_code == 200:
                data = response.json()
                return data.get('data', {}).get('markdown', '')  # No limit for Deep Research
                
        except Exception as e:
            print(f"   [WARNING] Firecrawl error: {e}")
        
        return ""
    
    def scrape_website_crawl4ai(self, url: str) -> str:
        """Step 2a: Scrape website using Crawl4AI (FREE, local browser)"""
        if not url:
            return ""
        
        try:
            from sources.crawl4ai_scraper import scrape_url
            
            content = scrape_url(url, timeout=30)
            
            if content:
                self.stats['crawl4ai_calls'] += 1
                return content  # No limit for Deep Research
                
        except Exception as e:
            print(f"   [WARNING] Crawl4AI error: {e}")
        
        return ""
    
    def scrape_website(self, url: str) -> str:
        """
        Unified scraping method that respects user's scraper choice.
        Falls back through: Crawl4AI -> Firecrawl -> HTTP
        """
        if not url:
            return ""
        
        content = ""
        
        if self.scraper_choice == "crawl4ai":
            # Crawl4AI only
            content = self.scrape_website_crawl4ai(url)
            if not content:
                print("   [FALLBACK] Crawl4AI failed, trying HTTP...")
                content = self.scrape_website_http(url)
                
        elif self.scraper_choice == "firecrawl":
            # Firecrawl only
            content = self.scrape_website_firecrawl(url)
            if not content:
                print("   [FALLBACK] Firecrawl failed, trying HTTP...")
                content = self.scrape_website_http(url)
                
        else:  # "both" or default
            # Crawl4AI first, then Firecrawl, then HTTP
            content = self.scrape_website_crawl4ai(url)
            if not content:
                print("   [FALLBACK] Crawl4AI failed, trying Firecrawl...")
                content = self.scrape_website_firecrawl(url)
            if not content:
                print("   [FALLBACK] Firecrawl failed, trying HTTP...")
                content = self.scrape_website_http(url)
        
        return content
    
    def scrape_website_http(self, url: str) -> str:
        """Step 2b: Fallback - simple HTTP scraping"""
        if not url:
            return ""
        
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            self.stats['http_calls'] += 1
            
            if response.status_code == 200:
                # Simple text extraction
                text = response.text
                # Remove scripts, styles
                text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL)
                text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
                text = re.sub(r'<[^>]+>', ' ', text)
                # Clean up whitespace
                text = ' '.join(text.split())
                return text
                
        except Exception as e:
            print(f"   [WARNING] HTTP error: {e}")
        
        return ""
    
    def extract_with_gemini(self, company_name: str, website: str, context: str, search_criteria: Dict = None) -> Dict:
        """Step 3: Use Gemini to extract structured data AND score based on real data (250 req/day)"""

        self._wait_for_gemini_rate_limit()

        # Get reference companies and criteria for scoring
        ref_companies = ""
        revenue_range = "$5M-$50M"
        target_industry = "packaging"

        if search_criteria:
            refs = search_criteria.get('reference_companies', [])
            if refs:
                ref_companies = f"Reference companies to compare against: {', '.join(refs[:3])}"

            rev = search_criteria.get('revenue', {})
            rev_min = rev.get('revenue_min_millions', 5)
            rev_max = rev.get('revenue_max_millions', 50)
            revenue_range = f"${rev_min}M-${rev_max}M"

            ind = search_criteria.get('industry', {})
            industries = ind.get('industry', ['packaging'])
            target_industry = ', '.join(industries[:3])

        prompt = f"""Extract M&A screening data and SCORE this company based on REAL data.

Company: {company_name}
Website: {website}
Additional Context: {context}

TARGET CRITERIA:
- Industry: {target_industry}
- Target Revenue Range: {revenue_range}
{ref_companies}

IMPORTANT: You MUST respond with ONLY valid JSON. No markdown, no explanation, just JSON.

Return this EXACT structure:
{{
  "name": "{company_name}",
  "website": "{website}",
  "year_established": <year as number or null>,
  "revenue_estimate": "<format: '$25M' or '<$5M' or '$5-10M' or '>$1B' or 'unknown'>",
  "employee_count": "<number or range like '50-200' or 'unknown'>",
  "line_of_business": "<B2B/B2C/Wholesale Distributor/etc>",
  "core_categories": "<main products>",
  "ownership": "<private/family-owned/PE-backed/public>",
  "sales_channel": "<online/retail/B2B distributors/hybrid>",
  "headquarters": "<city, state format>",
  "region": "US",
  "manufacturing": "<describe facilities or 'N/A' or 'unknown'>",
  "synergies": "<2-3 bullet points on strategic fit>",
  "fit_score": <0-100 based on REAL scoring rules below>,
  "score_reason": "<1 sentence explaining score>",
  "priority": "<High/Medium/Low based on fit_score>",
  "notes": "<any notable facts>",
  "status": "potential",
  "vendor_type": "potential"
}}

SCORING RULES (use fit_score field):

0-15: DISQUALIFIED
- Revenue > $1B (Fortune 500, global leaders)
- Public company with large market cap
- Wrong industry entirely
- Company no longer exists

15-35: POOR FIT
- Revenue $500M-$1B (too large)
- Right industry but wrong business model
- Has been acquired

35-55: MODERATE FIT
- Revenue unclear but seems mid-market
- Right industry, uncertain size
- Geography match unclear

55-75: GOOD FIT
- Revenue appears to be in target range
- Matches industry and geography
- Private/family-owned company

75-90: EXCELLENT FIT
- Revenue clearly in target range
- Direct competitor to reference companies
- Similar business model (e-commerce, custom, DTC)

90-100: PERFECT FIT
- Almost identical to reference companies
- Confirmed revenue in range
- Strong synergies identified

Set priority based on fit_score:
- High: fit_score >= 70
- Medium: fit_score 40-69
- Low: fit_score < 40

JSON ONLY:"""

        try:
            response = model.generate_content(prompt)
            self.stats['gemini_calls'] += 1
            
            json_text = response.text.strip()
            
            # Extract JSON from markdown if present
            if '```json' in json_text:
                json_text = json_text.split('```json')[1].split('```')[0].strip()
            elif '```' in json_text:
                json_text = json_text.split('```')[1].split('```')[0].strip()
            
            data = json.loads(json_text)
            return data
            
        except json.JSONDecodeError as e:
            print(f"   [ERROR] JSON parse error: {e}")
            print(f"   Raw response: {response.text[:200]}")
            return self._create_fallback_data(company_name, website)
            
        except Exception as e:
            print(f"   [ERROR] Porto error: {e}")
            return self._create_fallback_data(company_name, website)
    
    def _create_fallback_data(self, company_name: str, website: str) -> Dict:
        """Create minimal data structure if extraction fails"""
        return {
            "name": company_name,
            "website": website,
            "year_established": None,
            "revenue_estimate": "unknown",
            "line_of_business": "Promotional Products",
            "core_categories": "unknown",
            "ownership": "private",
            "sales_channel": "B2B",
            "headquarters": "unknown",
            "region": "US",
            "manufacturing": "unknown",
            "employees": "unknown",
            "synergies": "Promotional products capabilities",
            "priority": "Medium",
            "notes": "Data incomplete - requires manual review",
            "status": "potential",
            "vendor_type": "potential"
        }
    
    def enrich_company(self, company_name: str, search_criteria: Dict = None) -> Dict:
        """Full enrichment pipeline with smart fallbacks"""
        print(f"\n🔍 Enriching: {company_name}")

        # Step 1: Search for company
        search_data = self.search_company(company_name)
        website = search_data.get('website', '')
        snippet = search_data.get('snippet', '')

        if website:
            print(f"   [OK] Found: {website}")
        else:
            print(f"   [WARNING] No website found via search")

        # Step 2: Scrape website content (uses user's preferred scraper)
        content = ""
        if website:
            print(f"   -> Scraping website...")
            content = self.scrape_website(website)
            if content:
                print(f"   [OK] Got {len(content)} chars")

        # Combine all context
        full_context = f"{snippet} {content}".strip()

        # Step 3: Extract with Gemini AND score based on real data
        company_data = {}
        
        # CHECK: Use Deep Research? (Default to True if available for this demo, or add a toggle)
        use_deep = DEEP_RESEARCH_AVAILABLE # Can be toggled via search_criteria
        
        if use_deep and self.deep_enricher:
            try:
                print(f"   -> 🚀 Using DEEP RESEARCH Engine...")
                # Pass initial context (snippet + home page scrape) to Deep Research
                enriched_obj = self.deep_enricher.enrich_company(
                    company_name=company_name, 
                    domain=None, # Will be extracted if needed
                    website=website,
                    initial_context=full_context
                )
                
                # Convert Pydantic to Dict for legacy compatibility
                # We need to ensure the dict matches what the rest of the system expects
                company_data = enriched_obj.model_dump()
                
                # Normalize keys if needed (e.g. map new fields to old ones if missing)
                company_data['fit_score'] = company_data.get('ma_fit_score', 0)
                company_data['priority'] = company_data.get('priority', 'Medium')
                
                # Flatten research metadata if needed for simple display
                if company_data.get('research_metadata'):
                    company_data['research_grade'] = company_data['research_metadata'].get('research_grade')
                    company_data['hallucination_risk_score'] = company_data['research_metadata'].get('hallucination_risk_score')
                    
            except Exception as e:
                print(f"   [ERROR] Deep Research failed: {e}. Falling back to standard extraction.")
                print(f"   Traceback: {e}")
                use_deep = False
        
        if not use_deep:
            print(f"   -> Extracting & scoring (Standard Mode)...")
            company_data = self.extract_with_gemini(company_name, website, full_context, search_criteria)
        
        # Step 4: Try SEC EDGAR for verified financials (US public companies)
        sec_data = self._try_sec_validation(company_name)
        if sec_data and sec_data.get('found'):
            print(f"   [SEC] Verified: {sec_data.get('ticker', '')} - Revenue: ${sec_data.get('revenue_usd', 0)/1e6:.0f}M" if sec_data.get('revenue_usd') else f"   [SEC] Found: {sec_data.get('ticker', company_name)}")
            
            # Add verified data with confidence scores
            if sec_data.get('revenue_usd'):
                company_data['revenue_verified'] = f"${sec_data['revenue_usd']/1e6:.0f}M"
                company_data['revenue_year'] = sec_data.get('revenue_year', '')
                company_data['revenue_confidence'] = 'verified'
            
            company_data['ticker'] = sec_data.get('ticker', '')
            company_data['cik'] = sec_data.get('cik', '')
            company_data['sic_code'] = sec_data.get('sic', '')
            company_data['is_public'] = True
        else:
            company_data['revenue_confidence'] = 'estimated'
            company_data['is_public'] = False
        
        # Add source
        company_data['source'] = website or "Search required"

        # Show score in output
        fit_score = company_data.get('fit_score', 'N/A')
        priority = company_data.get('priority', 'Unknown')
        print(f"   [SUCCESS] Complete - Score: {fit_score}/100, Priority: {priority}")

        return company_data
    
    def _try_sec_validation(self, company_name: str) -> Optional[Dict]:
        """Try to get verified SEC data for a company."""
        try:
            from sources.sec_edgar import get_company_financials
            return get_company_financials(company_name)
        except Exception as e:
            return None

def get_project_root():
    """Get the project root directory (parent of enrichment/)"""
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def create_professional_excel(
    companies: List[Dict], 
    industry: str = "Unknown",
    reference_company: str = "",
    session_id: int = 0
):
    """
    Create Excel matching their EXACT format with professional styling.
    Filename format: [industry]_[reference_company]_session[N].xlsx
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M&A Target Screening"
    
    # Sort companies by fit_score (highest first), handling None values
    def safe_score(x):
        score = x.get('fit_score')
        if score is None:
            return -1  # Sort None scores to the bottom
        try:
            return float(score)
        except (ValueError, TypeError):
            return -1
    
    companies = sorted(companies, key=safe_score, reverse=True)

    # Headers with fit_score added
    headers = [
        "No.",
        "Fit Score",
        "Source",
        "Name",
        "Website",
        "Year of Est.",
        "Size (Revenues USD mn)",
        "Verified Revenue",  # NEW
        "Research Grade",    # NEW
        "Risk Score",        # NEW
        "LOB",
        "Core categories",
        "Ownership",
        "channel of sales",
        "hocity",
        "Region",
        "Manufacturing",
        "Employees",
        "Areas of Synergies",
        "Priority",
        "Score Reason",
        "Notes",
        "Status",
        "Vendor"
    ]
    
    # Header styling - professional dark blue
    ws.append(headers)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style='thin', color='000000')
        )
    
    # Add data rows
    for idx, company in enumerate(companies, 1):
        # Helper function to convert values to Excel-safe strings
        def to_excel_value(val):
            if isinstance(val, list):
                return '\n'.join(str(v) for v in val)
            elif isinstance(val, dict):
                return str(val)
            else:
                return val
        
        row_data = [
            idx,
            to_excel_value(company.get('fit_score') or company.get('ma_fit_score', '')),
            to_excel_value(company.get('source', '')),
            to_excel_value(company.get('name') or company.get('company_name', '')),
            to_excel_value(company.get('website', '')),
            to_excel_value(company.get('year_established', '')),
            to_excel_value(company.get('revenue_estimate', '')),
            to_excel_value(company.get('revenue_verified', company.get('verified_revenue', ''))), # NEW
            to_excel_value(company.get('research_grade', 'N/A')), # NEW
            to_excel_value(company.get('hallucination_risk_score', '')), # NEW
            to_excel_value(company.get('line_of_business', '')),
            to_excel_value(company.get('core_categories', '')),
            to_excel_value(company.get('ownership', '')),
            to_excel_value(company.get('sales_channel', '')),
            to_excel_value(company.get('headquarters', '')),
            to_excel_value(company.get('region', 'US')),
            to_excel_value(company.get('manufacturing', '')),
            to_excel_value(company.get('employee_count', company.get('employees', ''))),
            to_excel_value(company.get('synergies', '')),
            to_excel_value(company.get('priority', 'Medium')),
            to_excel_value(company.get('score_reason', '')),
            to_excel_value(company.get('notes', '')),
            to_excel_value(company.get('status', 'potential')),
            to_excel_value(company.get('vendor_type', 'potential'))
        ]
        ws.append(row_data)
        
        # Priority color coding
        priority = company.get('priority', 'Medium')
        priority_colors = {
            'High': 'C6EFCE',    # Light green
            'Medium': 'FFEB9C',  # Light yellow
            'Low': 'FFC7CE'      # Light red
        }
        
        row_num = idx + 1
        
        # Color the priority cell (column 17 now)
        if priority in priority_colors:
            priority_cell = ws.cell(row=row_num, column=17)
            priority_cell.fill = PatternFill(
                start_color=priority_colors[priority],
                end_color=priority_colors[priority],
                fill_type="solid"
            )
            priority_cell.font = Font(bold=True)

        # Color fit_score cell based on score (column 2)
        fit_score = company.get('fit_score') or company.get('ma_fit_score')
        if fit_score is None:
            fit_score = 0
        # Ensure fit_score is numeric for comparison
        try:
            fit_score = float(fit_score)
        except (TypeError, ValueError):
            fit_score = 0
        score_cell = ws.cell(row=row_num, column=2)
        if fit_score >= 70:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif fit_score >= 40:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        score_cell.font = Font(bold=True)

        # Light alternating row colors for readability
        if idx % 2 == 0:
            for col in range(1, 25): # Increased column count
                if col not in [2, 17]:  # Skip score and priority cells
                    ws.cell(row=row_num, column=col).fill = PatternFill(
                        start_color="F2F2F2",
                        end_color="F2F2F2",
                        fill_type="solid"
                    )
    
    # Auto-size columns
    column_widths = {
        'A': 6,   # No.
        'B': 10,  # Fit Score
        'C': 45,  # Source
        'D': 25,  # Name
        'E': 35,  # Website
        'F': 12,  # Year
        'G': 20,  # Revenue
        'H': 20,  # LOB
        'I': 30,  # Core categories
        'J': 18,  # Ownership
        'K': 18,  # Sales channel
        'L': 25,  # HQ
        'M': 10,  # Region
        'N': 30,  # Manufacturing
        'O': 15,  # Employees
        'P': 40,  # Synergies
        'Q': 12,  # Priority
        'R': 45,  # Score Reason
        'S': 40,  # Notes
        'T': 12,  # Status
        'U': 12,  # Vendor
        'V': 15,  # Verified Revenue
        'W': 10,  # Grade
        'X': 10   # Risk
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Ensure exports directory exists and save there
    project_root = get_project_root()
    exports_dir = os.path.join(project_root, 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    
    # Generate dynamic filename: [industry]_[reference_company]_session[N].xlsx
    import re
    def sanitize_name(name):
        return re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')[:30]
    
    industry_safe = sanitize_name(industry) if industry else "Unknown"
    ref_safe = sanitize_name(reference_company) if reference_company else "NoRef"
    session_str = f"session{session_id}" if session_id else "all"
    
    output_file = f"{industry_safe}_{ref_safe}_{session_str}.xlsx"
    output_path = os.path.join(exports_dir, output_file)
    
    # Save
    wb.save(output_path)
    print(f"\n[SUCCESS] Excel created: exports/{output_file}")


def main():
    """Main demo execution"""
    
    print("=" * 80)
    print("=" * 80)
    print(" M&A SCREENER - DEMO")
    print("=" * 80)
    print("[TIME] ~Est. 3 minutes per company")
    print()
    
    # Check API keys
    required_keys = {
        'GEMINI_API_KEY': 'Gemini (REQUIRED)',
        'SERPER_KEY': 'Serper (recommended)',
        'FIRECRAWL_KEY': 'Firecrawl (optional)'
    }
    
    print(" API Key Check:")
    for key, name in required_keys.items():
        status = "[OK]" if os.getenv(key) else "[MISSING]"
        print(f"   {status} {name}")
    
    if not os.getenv('GEMINI_API_KEY'):
        print("\n[ERROR] GEMINI_API_KEY is required!")
        print("   Get free key: https://aistudio.google.com/app/apikey")
        return
    
    # ─────────────────────────────────────────────────────────────
    # LOAD COMPANIES FROM DATABASE (with smart filtering)
    # ─────────────────────────────────────────────────────────────
    target_companies = []
    company_data_map = {}  # name -> full company data
    
    # Track session metadata for Excel filename
    selected_session_id = None
    selected_industry = "Unknown"
    selected_reference = ""
    
    try:
        from data.db import get_db_connection, get_database_stats, get_all_sessions, get_companies_by_session, DB_PATH
        import os as os_module
        
        # Check if database exists
        if os_module.path.exists(DB_PATH):
            stats = get_database_stats()
            total = stats.get('unique_companies', 0)
            enriched = stats.get('enriched', 0)
            discovered = stats.get('discovered', 0)
            
            print(f"\n[DB] Database Status:")
            print(f"     Total companies: {total}")
            print(f"     Already enriched: {enriched}")
            print(f"     Need enrichment: {discovered}")
            
            if total == 0:
                print("\n[WARNING] No companies in database!")
                print("[INFO] Run 'python run_discovery.py' first")
                return
            
            # ─────────────────────────────────────────────────────────────
            # SESSION SELECTION (STEP 1)
            # ─────────────────────────────────────────────────────────────
            sessions = get_all_sessions()
            
            if len(sessions) > 1:
                print("\n" + "─"*50)
                print("SESSION SELECTION")
                print("─"*50)
                print("\n[DB] Available Sessions:")
                
                for i, s in enumerate(sessions, 1):
                    date = s['created_at'][:10] if s['created_at'] else 'Unknown'
                    unenriched = s['discovered_count']
                    total_in_session = s['company_count']
                    industry = s['industry']
                    print(f"     [{i}] Session #{s['id']} ({date}) - {total_in_session} companies - \"{industry}\" ({unenriched} need enrichment)")
                
                print(f"     [{len(sessions)+1}] All sessions combined - {total} companies")
                print(f"     [{len(sessions)+2}] Latest session only (#{sessions[0]['id']})")
                
                session_choice = input(f"\nSelect session [1-{len(sessions)+2}]: ").strip()
                
                try:
                    session_idx = int(session_choice) - 1
                    
                    if session_idx == len(sessions):
                        # All sessions
                        selected_session_id = None
                        print(f"\n[OK] Using all sessions ({total} companies)")
                    elif session_idx == len(sessions) + 1:
                        # Latest session
                        selected_session_id = sessions[0]['id']
                        selected_industry = sessions[0].get('industry', 'Unknown')
                        print(f"\n[OK] Using latest session #{selected_session_id} - {selected_industry}")
                    elif 0 <= session_idx < len(sessions):
                        # Specific session
                        selected_session_id = sessions[session_idx]['id']
                        selected_industry = sessions[session_idx].get('industry', 'Unknown')
                        print(f"\n[OK] Using session #{selected_session_id} - {selected_industry}")
                    else:
                        selected_session_id = None
                        print("\n[WARNING] Invalid, using all sessions")
                except:
                    selected_session_id = None
                    print("\n[WARNING] Invalid, using all sessions")
            else:
                # Only one session, use it automatically
                selected_session_id = sessions[0]['id'] if sessions else None
                if sessions:
                    print(f"\n[INFO] Using session #{selected_session_id} - {sessions[0]['industry']}")
            
            # ─────────────────────────────────────────────────────────────
            # LOAD COMPANIES (from selected session or all)
            # ─────────────────────────────────────────────────────────────
            if selected_session_id:
                all_companies = get_companies_by_session(selected_session_id, status='discovered')
            else:
                # Load from all sessions
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT name, domain, website, score, status, raw_data_json
                        FROM companies 
                        WHERE status != 'enriched'
                        ORDER BY score DESC NULLS LAST
                    """)
                    all_companies = [dict(row) for row in cursor.fetchall()]
            
            # Check if there are companies to enrich
            print(f"\n[INFO] Found {len(all_companies)} companies needing enrichment")
            
            if len(all_companies) == 0:
                print("\n[SUCCESS] All companies already enriched!")
                print("[HINT] To re-enrich, select option [4] or run discovery again")
                return
            
            # ─────────────────────────────────────────────────────────────
            # FILTERING OPTIONS (STEP 2)
            # ─────────────────────────────────────────────────────────────
            print("\n" + "─"*50)
            print("COMPANY SELECTION")
            print("─"*50)
            print(f"\n[1] All {len(all_companies)} companies needing enrichment")
            print("[2] By minimum fit score (e.g., only score > 70)")
            print("[3] By count (e.g., top 20 by score)")
            print("[4] Refresh all (re-enrich everything, including already done)")
            
            choice = input("\nSelect option [1/2/3/4]: ").strip()
            
            if choice == "1":
                # All companies needing enrichment
                selected = all_companies
                print(f"\n[OK] Will enrich all {len(selected)} companies")
                
            elif choice == "2":
                # By minimum score
                min_score = input("Enter minimum fit score (0-100): ").strip()
                try:
                    min_score = int(min_score)
                except:
                    min_score = 70
                    print(f"[WARNING] Invalid, using {min_score}")
                
                selected = [c for c in all_companies if (c.get('score') or 0) >= min_score]
                print(f"\n[OK] Found {len(selected)} companies with score >= {min_score}")
                
                if len(selected) == 0:
                    print("[WARNING] No companies match that criteria!")
                    return
                    
            elif choice == "3":
                # By count
                count = input("Enter number of companies (default 15): ").strip()
                try:
                    count = int(count) if count else 15
                except:
                    count = 15
                
                selected = all_companies[:count]
                print(f"\n[OK] Will enrich top {len(selected)} companies by score")
                
            elif choice == "4":
                # Refresh all - include enriched
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT name, domain, website, score, status, raw_data_json
                        FROM companies 
                        ORDER BY score DESC NULLS LAST
                    """)
                    selected = [dict(row) for row in cursor.fetchall()]
                
                print(f"\n[OK] Will re-enrich ALL {len(selected)} companies")
                
            else:
                print(f"[WARNING] Invalid choice, using all {len(all_companies)}")
                selected = all_companies
            
            # Build target list
            for c in selected:
                name = c.get('name', '').strip()
                if name:
                    target_companies.append(name)
                    company_data_map[name] = c
            
            print(f"\n[INFO] Selected {len(target_companies)} companies for enrichment\n")
            
        else:
            print("[WARNING] No database found!")
            print("[INFO] Falling back to JSON file...")
            raise Exception("No database")
            
    except Exception as e:
        # Fallback to JSON if database fails
        print(f"[WARNING] Database error: {e}")
        print("[INFO] Falling back to discovered_companies.json...")
        
        project_root = get_project_root()
        discovered_path = os.path.join(project_root, 'output', 'discovered_companies.json')
        
        if os.path.exists(discovered_path):
            with open(discovered_path, 'r') as f:
                data = json.load(f)
                discovered = data.get('companies', [])
                discovered.sort(key=lambda x: x.get('fit_score', 0), reverse=True)
                
                num_to_enrich = input(f"\nFound {len(discovered)} companies. How many to enrich? (default 15): ").strip()
                num_to_enrich = int(num_to_enrich) if num_to_enrich.isdigit() else 15
                
                target_companies = [c['name'] for c in discovered[:num_to_enrich]]
        else:
            print("[ERROR] No companies found! Run discovery first.")
            return
    
    print(f"[INFO] Processing {len(target_companies)} companies...\n")

    # Load search criteria for proper scoring
    search_criteria = None
    criteria_path = os.path.join(get_project_root(), 'output', 'search_criteria.json')
    if os.path.exists(criteria_path):
        try:
            with open(criteria_path, 'r') as f:
                data = json.load(f)
                search_criteria = data.get('criteria', {})
                refs = search_criteria.get('reference_companies', [])
                if refs:
                    selected_reference = refs[0] if refs else ""
                    print(f"[INFO] Scoring against reference companies: {', '.join(refs[:3])}")
                
                # Get industry from criteria if not already set
                if selected_industry == "Unknown":
                    ind = search_criteria.get('industry', {})
                    industries = ind.get('industry', [])
                    if industries:
                        selected_industry = industries[0]
        except Exception as e:
            print(f"[WARNING] Could not load search criteria: {e}")

    enricher = MultiSourceEnricher()

    # Ask user which scraper to use
    enricher.ask_scraper_preference()

    enriched_companies = []

    start_time = time.time()

    for company_name in target_companies:
        try:
            company_data = enricher.enrich_company(company_name, search_criteria)
            if company_data:
                enriched_companies.append(company_data)
        except Exception as e:
            print(f"[ERROR] Error with {company_name}: {e}")
            continue
    
    elapsed = time.time() - start_time
    
    # ─────────────────────────────────────────────────────────────
    # SAVE ENRICHMENT TO DATABASE
    # ─────────────────────────────────────────────────────────────
    try:
        from data.db import get_db_connection, DB_PATH
        import os as os_module
        
        if os_module.path.exists(DB_PATH):
            with get_db_connection() as conn:
                cursor = conn.cursor()
                saved_count = 0
                
                for company in enriched_companies:
                    try:
                        # Handle different key names from different enrichment sources:
                        # - extract_with_gemini uses 'name'
                        # - Deep Research (EnrichedCompany) uses 'company_name'
                        name = company.get('name') or company.get('company_name') or company.get('company', '')
                        enriched_json = json.dumps(company)
                        
                        # Update company with enriched data and change status
                        cursor.execute("""
                            UPDATE companies 
                            SET status = 'enriched',
                                enriched_at = datetime('now'),
                                enriched_data_json = ?
                            WHERE LOWER(name) = LOWER(?)
                        """, (enriched_json, name))
                        
                        if cursor.rowcount > 0:
                            saved_count += 1
                            
                    except Exception as e:
                        print(f"   [WARNING] Failed to save {name}: {e}")
                
                print(f"\n[DB] Saved {saved_count} enriched companies to database")
                
    except Exception as e:
        print(f"[WARNING] Database save failed: {e}")
    
    # Create Excel
    print(f"\n[INFO] Creating Excel file...")
    create_professional_excel(
        enriched_companies,
        industry=selected_industry,
        reference_company=selected_reference,
        session_id=selected_session_id or 0
    )
    
    # Summary
    print("\n" + "=" * 80)
    print(" ENRICHMENT SUMMARY")
    print("=" * 80)
    print(f"[SUCCESS] Companies processed: {len(enriched_companies)}/{len(target_companies)}")
    print(f"[TIME] Total time: {elapsed/60:.1f} minutes")
    if len(enriched_companies) > 0:
        print(f"[AVG] Avg per company: {elapsed/len(enriched_companies):.1f} seconds")
    print()
    
    # Priority breakdown
    priority_counts = {}
    for c in enriched_companies:
        p = c.get('priority', 'Unknown')
        priority_counts[p] = priority_counts.get(p, 0) + 1
    
    print(" Priority Distribution:")
    for priority, count in sorted(priority_counts.items()):
        print(f"   {priority}: {count} companies")
    
    print()
    print(" API Usage (LOCAL):")
    
    # Aggregate stats from deep enricher if used
    if enricher.deep_enricher:
        deep_stats = enricher.deep_enricher.get_stats()
        enricher.stats['grounding_calls'] += deep_stats.get('grounding_calls', 0)
        enricher.stats['gemini_calls'] += deep_stats.get('gemini_calls', 0)
    
    for service, count in enricher.stats.items():
        print(f"   {service}: {count} calls")
    
    print()
    print("[COST] Total Cost: $0")
    print()
    print("=" * 80)
    print(" Enrichment complete!")
    print(f" - Excel: exports/{selected_industry}_{selected_reference}_session{selected_session_id or 0}.xlsx")
    print(" - Database: All companies marked as 'enriched'")
    print("=" * 80)


if __name__ == "__main__":
    main()
